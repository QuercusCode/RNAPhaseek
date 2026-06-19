"""
Generic supplementary-table parser for LLPS literature gene lists.

Many LLPS papers post their gene lists as supplementary xlsx, csv, or tsv files.
Whoever downloads them (manually from journal websites or via the literature
workflow) drops them into Data/raw/multispecies/papers/ — this module finds
gene-identifier columns automatically and extracts deduplicated lists.

Usage (programmatic):
  from Functions.data_collection_multispecies.parse_supplementary import (
      parse_table, extract_gene_ids,
  )
  ids = extract_gene_ids("Data/raw/multispecies/papers/hubstenberger_2017.xlsx",
                          species="yeast")

Usage (CLI):
  python -m Functions.data_collection_multispecies.parse_supplementary \\
      --file Data/raw/multispecies/papers/lee_2020.xlsx \\
      --species celegans
"""

import argparse
import os
import re
import sys
from pathlib import Path
from typing import Iterable


# Patterns matching common ID-column headers per species
ID_COLUMN_PATTERNS = {
    "yeast": [
        r"orf", r"systematic", r"sgd", r"locus", r"gene[\s_]*(?:id|name)?", r"feature",
    ],
    "celegans": [
        r"wbgene", r"wormbase", r"gene[\s_]*id", r"sequence[\s_]*name", r"public[\s_]*name",
    ],
    "drosophila": [
        r"fbgn", r"flybase", r"gene[\s_]*id", r"symbol", r"cg[\s_]*number",
    ],
    "mouse": [
        r"ensmusg", r"ensembl", r"gene[\s_]*id", r"symbol", r"mgi",
    ],
    "human": [
        r"ensg", r"ensembl", r"gene[\s_]*id", r"symbol", r"hgnc",
    ],
    "arabidopsis": [
        r"at[a-z]?g\d+", r"tair", r"gene[\s_]*id", r"locus",
    ],
}

# Patterns matching ID VALUES (used to score columns when headers are ambiguous)
ID_VALUE_PATTERNS = {
    "yeast":       re.compile(r"^Y[A-P][LR]\d{3}[WC](?:-[A-Z])?$"),      # systematic ORF
    "celegans":    re.compile(r"^WBGene\d{8}$|^[A-Z0-9]+\.\d+[a-z]?$"),  # WBGene or transcript
    "drosophila":  re.compile(r"^FBgn\d{7}$|^CG\d+$"),                    # FBgn or CG number
    "mouse":       re.compile(r"^ENSMUSG\d{11}(?:\.\d+)?$"),
    "human":       re.compile(r"^ENSG\d{11}(?:\.\d+)?$"),
    "arabidopsis": re.compile(r"^AT[1-5MC]G\d{5}(?:\.\d+)?$"),
}


def parse_table(path: str | Path) -> list[dict]:
    """Read an xlsx / csv / tsv into a list of row dicts (header-aware)."""
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xls"):
        try:
            import openpyxl
        except ImportError:
            raise ImportError("pip install openpyxl")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        rows = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            srows = list(ws.iter_rows(values_only=True))
            if not srows:
                continue
            # Find the actual header row — first row with >=3 non-empty cells
            # that contains a string column name resembling an ID
            header_idx = 0
            for i, row in enumerate(srows[:10]):
                ne = [str(c).strip() for c in row if c is not None and str(c).strip()]
                if len(ne) >= 3:
                    header_idx = i; break
            header = [str(c).strip() if c is not None else f"col_{j}"
                      for j, c in enumerate(srows[header_idx])]
            for row in srows[header_idx + 1:]:
                d = {h: row[j] if j < len(row) else None for j, h in enumerate(header)}
                d["__sheet__"] = sheet
                rows.append(d)
        return rows
    if suffix == ".csv":
        import csv
        with open(path, encoding="utf-8") as f:
            return list(csv.DictReader(f))
    if suffix in (".tsv", ".txt"):
        import csv
        with open(path, encoding="utf-8") as f:
            return list(csv.DictReader(f, delimiter="\t"))
    raise ValueError(f"Unsupported file type: {suffix}")


def _score_column(values: list, species: str) -> float:
    """
    Heuristic score for whether a column contains gene IDs for this species.
    Score = fraction of non-empty values that match the species' ID_VALUE_PATTERN.
    """
    pat = ID_VALUE_PATTERNS.get(species)
    if pat is None:
        return 0.0
    non_empty = [str(v).strip() for v in values if v is not None and str(v).strip()]
    if not non_empty:
        return 0.0
    return sum(1 for v in non_empty if pat.match(v)) / len(non_empty)


def find_id_column(rows: list[dict], species: str) -> str | None:
    """Identify the best gene-ID column by header name + value patterns."""
    if not rows:
        return None
    cols = [c for c in rows[0].keys() if c != "__sheet__"]
    if not cols:
        return None

    name_pats = [re.compile(p, re.I) for p in ID_COLUMN_PATTERNS.get(species, [])]

    best, best_score = None, -1.0
    for c in cols:
        name_hit = any(p.search(str(c)) for p in name_pats)
        values = [r.get(c) for r in rows[:200]]
        value_score = _score_column(values, species)
        score = (1.0 if name_hit else 0.0) + value_score
        if score > best_score:
            best, best_score = c, score
    return best


def extract_gene_ids(path: str | Path, species: str,
                     id_column: str | None = None,
                     id_column_pattern: str | None = None,
                     ) -> list[str]:
    """
    Extract gene IDs from a supplementary table.

    species: 'yeast' | 'celegans' | 'drosophila' | 'mouse' | 'human' | 'arabidopsis'
    id_column: exact column header (overrides auto-detection)
    id_column_pattern: regex on column header to find the right column
    """
    rows = parse_table(path)
    if not rows:
        return []
    if id_column is None:
        if id_column_pattern:
            p = re.compile(id_column_pattern, re.I)
            for c in rows[0].keys():
                if c != "__sheet__" and p.search(str(c)):
                    id_column = c; break
        if id_column is None:
            id_column = find_id_column(rows, species)
    if id_column is None:
        raise ValueError(
            f"Could not identify ID column for {species}. "
            f"Available columns: {[c for c in rows[0].keys() if c != '__sheet__']}"
        )

    seen, ids = set(), []
    pat = ID_VALUE_PATTERNS.get(species)
    for r in rows:
        v = r.get(id_column)
        if v is None: continue
        v = str(v).strip().split(".")[0]
        if not v: continue
        if pat and not pat.match(v):
            # Accept symbols only when no strict pattern available
            continue
        if v not in seen:
            seen.add(v); ids.append(v)
    return ids


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--file",    required=True)
    p.add_argument("--species", required=True,
                   choices=list(ID_COLUMN_PATTERNS.keys()))
    p.add_argument("--id_column", default=None, help="Force exact column name")
    p.add_argument("--id_column_pattern", default=None, help="Regex to find column name")
    args = p.parse_args()
    ids = extract_gene_ids(args.file, args.species,
                           id_column=args.id_column,
                           id_column_pattern=args.id_column_pattern)
    print(f"Extracted {len(ids)} unique IDs from {args.file}")
    for x in ids[:25]:
        print(f"  {x}")
    if len(ids) > 25:
        print(f"  ... ({len(ids) - 25} more)")


if __name__ == "__main__":
    main()
